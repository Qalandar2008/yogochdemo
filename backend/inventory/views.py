import traceback
from decimal import Decimal, InvalidOperation
from datetime import datetime
from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import authenticate
from django.contrib.auth import get_user_model
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import FileResponse
from io import BytesIO

from .models import Product, ProductSize, InventoryTransaction, DebtRecord
from .serializers import (
    ProductSerializer,
    ProductListSerializer,
    StatsSerializer,
    InventoryTransactionSerializer,
    DebtRecordSerializer,
)


def parse_occurred_at(value):
    if not value:
        return timezone.now()
    dt = parse_datetime(value)
    if dt is None:
        raise ValueError('Invalid datetime format.')
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.get_current_timezone())
    return dt


def build_stats_data():
    products = Product.objects.prefetch_related('sizes').all()
    total_products = products.count()
    total_quantity = sum(p.quantity for p in products)
    sold_transactions = InventoryTransaction.objects.filter(
        transaction_type=InventoryTransaction.TransactionType.OUT
    ).select_related('product')
    total_sold_quantity = sum(t.quantity for t in sold_transactions)
    total_sold_volume = sum(p.sold_volume for p in products)
    total_volume = sum(p.total_volume for p in products)
    total_inventory_value = sum(float(p.purchase_price) * p.quantity for p in products)
    total_revenue = sum(float(t.unit_price) * t.quantity for t in sold_transactions)
    total_profit = sum(float(t.unit_price - t.unit_cost) * t.quantity for t in sold_transactions)

    data = {
        'total_products': total_products,
        'total_quantity': total_quantity,
        'total_sold_quantity': total_sold_quantity,
        'total_sold_volume': round(total_sold_volume, 3),
        'total_volume': round(total_volume, 3),
        'total_inventory_value': round(total_inventory_value, 2),
        'total_revenue': round(total_revenue, 2),
        'total_profit': round(total_profit, 2),
    }
    serializer = StatsSerializer(data=data)
    serializer.is_valid(raise_exception=True)
    return serializer.validated_data, products


class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.prefetch_related('sizes').all()
    serializer_class = ProductSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None  # Disable pagination for simple array response
    
    def get_serializer_class(self):
        if self.action == 'list':
            return ProductListSerializer
        return ProductSerializer
    
    @action(detail=False, methods=['get'], url_path='stats')
    def stats(self, request):
        """Dashboard uchun statistika"""
        try:
            data, _ = build_stats_data()
            return Response(data)
        except Exception as e:
            print(f"Stats error: {str(e)}")
            print(traceback.format_exc())
            return Response(
                {'error': f'Statistics calculation error: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'], url_path='export/excel')
    def export_excel(self, request):
        """Excel hisobot yuklash"""
        try:
            stats_data, products = build_stats_data()
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Styles
            header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Sheet 1: Statistics
            ws1 = wb.create_sheet('Statistics')
            ws1.append(['Yog\'och Inventory - Statistics'])
            ws1.merge_cells('A1:C1')
            ws1['A1'].font = Font(bold=True, size=16)
            ws1['A1'].alignment = Alignment(horizontal='center')
            ws1.append([])
            
            # Headers
            ws1.append(['Ko\'rsatkich', 'Qiymat', 'Birlik'])
            for cell in ws1[3]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            
            # Data
            stats_rows = [
                ['Jami mahsulotlar', stats_data['total_products'], 'ta'],
                ['Ombordagi dona', stats_data['total_quantity'], 'dona'],
                ['Sotilgan dona', stats_data['total_sold_quantity'], 'dona'],
                ['Jami hajm', stats_data['total_volume'], 'm³'],
                ['Sotilgan hajm', stats_data['total_sold_volume'], 'm³'],
                ['Ombor qiymati', stats_data['total_inventory_value'], 'so\'m'],
                ['Jami tushum', stats_data['total_revenue'], 'so\'m'],
                ['Sof foyda', stats_data['total_profit'], 'so\'m'],
            ]
            
            for row in stats_rows:
                ws1.append(row)
                for cell in ws1[ws1.max_row]:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='left' if cell.column == 1 else 'right')
            
            # Adjust column widths
            ws1.column_dimensions['A'].width = 25
            ws1.column_dimensions['B'].width = 15
            ws1.column_dimensions['C'].width = 10
            
            # Sheet 2: Products
            ws2 = wb.create_sheet('Products')
            ws2.append(['Mahsulotlar Ro\'yxati'])
            ws2.merge_cells('A1:I1')
            ws2['A1'].font = Font(bold=True, size=16)
            ws2['A1'].alignment = Alignment(horizontal='center')
            ws2.append([])
            
            # Product headers
            product_headers = ['ID', 'Nomi', 'Dona', 'Narxi', 'O\'lcham', 'Umumiy hajm', 'Sotilgan', 'Sotilgan hajm', 'Foyda']
            ws2.append(product_headers)
            for cell in ws2[3]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            
            # Product data
            for product in products:
                sizes_str = ', '.join([f"{s.length}×{s.width}×{s.height}" for s in product.sizes.all()])
                ws2.append([
                    product.id,
                    product.name,
                    product.quantity,
                    float(product.purchase_price),
                    sizes_str,
                    round(product.total_volume, 3),
                    product.sold_quantity,
                    round(product.sold_volume, 3),
                    round(product.profit, 2)
                ])
                for cell in ws2[ws2.max_row]:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center' if cell.column in [1, 3, 7] else 'left')
            
            # Adjust column widths
            ws2.column_dimensions['A'].width = 5
            ws2.column_dimensions['B'].width = 25
            ws2.column_dimensions['C'].width = 10
            ws2.column_dimensions['D'].width = 12
            ws2.column_dimensions['E'].width = 20
            ws2.column_dimensions['F'].width = 12
            ws2.column_dimensions['G'].width = 12
            ws2.column_dimensions['H'].width = 14
            ws2.column_dimensions['I'].width = 12

            # Sheet 3: Summary
            ws3 = wb.create_sheet('Summary')
            ws3.append(['Metric', 'Value'])
            for cell in ws3[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            summary_rows = [
                ['Total Products', stats_data['total_products']],
                ['Total Quantity', stats_data['total_quantity']],
                ['Total Sold Quantity', stats_data['total_sold_quantity']],
                ['Total Volume (m3)', stats_data['total_volume']],
                ['Total Sold Volume (m3)', stats_data['total_sold_volume']],
                ['Total Inventory Value', stats_data['total_inventory_value']],
                ['Total Revenue', stats_data['total_revenue']],
                ['Total Profit', stats_data['total_profit']],
            ]
            for row in summary_rows:
                ws3.append(row)
                for cell in ws3[ws3.max_row]:
                    cell.border = thin_border
            ws3.column_dimensions['A'].width = 28
            ws3.column_dimensions['B'].width = 18
            
            # Save to buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            # Return file response
            return FileResponse(
                buffer,
                as_attachment=True,
                filename='inventory_report.xlsx',
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        except Exception as e:
            print(f"Excel export error: {str(e)}")
            print(traceback.format_exc())
            return Response(
                {'error': f'Excel export error: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'], url_path='restock')
    def restock(self, request, pk=None):
        product = self.get_object()
        quantity = request.data.get('quantity')
        purchase_price = request.data.get('purchase_price')

        try:
            quantity = int(quantity)
            if quantity <= 0:
                raise ValueError
        except (TypeError, ValueError):
            return Response(
                {'error': 'Quantity must be a positive integer.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            purchase_price = Decimal(str(purchase_price))
            if purchase_price <= 0:
                raise ValueError
        except (TypeError, ValueError, InvalidOperation):
            return Response(
                {'error': 'Purchase price must be a positive number.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            occurred_at = parse_occurred_at(request.data.get('occurred_at'))
        except ValueError:
            return Response({'error': 'Invalid occurred_at datetime format.'}, status=status.HTTP_400_BAD_REQUEST)

        product.quantity += quantity
        if product.quantity == 0:
            product.purchase_price = purchase_price
        else:
            current_total = Decimal(product.quantity) * product.purchase_price
            incoming_total = Decimal(quantity) * purchase_price
            product.purchase_price = (current_total + incoming_total) / Decimal(product.quantity + quantity)
        product.sold_quantity = max(product.sold_quantity, 0)
        product.save(update_fields=['quantity', 'purchase_price', 'updated_at'])
        InventoryTransaction.objects.create(
            product=product,
            transaction_type=InventoryTransaction.TransactionType.IN,
            quantity=quantity,
            unit_price=purchase_price,
            unit_cost=purchase_price,
            occurred_at=occurred_at,
        )

        serializer = ProductSerializer(product, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='sell')
    def sell(self, request, pk=None):
        product = self.get_object()
        quantity = request.data.get('quantity')
        sale_price = request.data.get('sale_price')

        try:
            quantity = int(quantity)
            if quantity <= 0:
                raise ValueError
        except (TypeError, ValueError):
            return Response(
                {'error': 'Quantity must be a positive integer.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            sale_price = Decimal(str(sale_price))
            if sale_price <= 0:
                raise ValueError
        except (TypeError, ValueError, InvalidOperation):
            return Response(
                {'error': 'Sale price must be a positive number.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            occurred_at = parse_occurred_at(request.data.get('occurred_at'))
        except ValueError:
            return Response({'error': 'Invalid occurred_at datetime format.'}, status=status.HTTP_400_BAD_REQUEST)

        if quantity > product.quantity:
            return Response(
                {'error': 'Cannot sell more than available stock.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        product_cost = product.purchase_price
        product.quantity -= quantity
        product.sold_quantity += quantity
        product.save(update_fields=['quantity', 'sold_quantity', 'updated_at'])
        InventoryTransaction.objects.create(
            product=product,
            transaction_type=InventoryTransaction.TransactionType.OUT,
            quantity=quantity,
            unit_price=sale_price,
            unit_cost=product_cost,
            occurred_at=occurred_at,
        )

        serializer = ProductSerializer(product, context={'request': request})
        response_data = serializer.data
        response_data['last_sale_price'] = sale_price
        return Response(response_data, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def stats_view(request):
    try:
        data, _ = build_stats_data()
        return Response(data)
    except Exception as e:
        print(f"Stats API error: {str(e)}")
        print(traceback.format_exc())
        return Response({'error': f'Statistics calculation error: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_excel_view(request):
    viewset = ProductViewSet()
    viewset.request = request
    return viewset.export_excel(request)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def transactions_view(request):
    transactions = InventoryTransaction.objects.select_related('product').order_by('-occurred_at', '-created_at')
    serializer = InventoryTransactionSerializer(transactions, many=True)
    return Response(serializer.data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_superuser_view(request):
    if not request.user.is_superuser:
        return Response({'error': 'Only superuser can create another superuser.'}, status=status.HTTP_403_FORBIDDEN)

    username = request.data.get('username')
    password = request.data.get('password')

    if not username or not password:
        return Response({'error': 'Username and password are required.'}, status=status.HTTP_400_BAD_REQUEST)

    user_model = get_user_model()
    if user_model.objects.filter(username=username).exists():
        return Response({'error': 'Username already exists.'}, status=status.HTTP_400_BAD_REQUEST)

    user_model.objects.create_superuser(username=username, email='', password=password)
    return Response({'message': 'Superuser created successfully.'}, status=status.HTTP_201_CREATED)


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def debts_view(request):
    if request.method == 'GET':
        records = DebtRecord.objects.select_related('product').order_by('-occurred_at')
        serializer = DebtRecordSerializer(records, many=True)
        return Response(serializer.data)

    serializer = DebtRecordSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    validated = serializer.validated_data

    product = validated['product']
    quantity = validated['quantity']
    unit_price = validated['unit_price']
    phone = validated['phone']

    if not phone.startswith('+998') or len(phone) != 13 or not phone[1:].isdigit():
        return Response({'error': 'Telefon raqam +998XXXXXXXXX formatda bo\'lishi kerak.'}, status=status.HTTP_400_BAD_REQUEST)

    if quantity > product.quantity:
        return Response({'error': 'Qarzga berish uchun ombordagi mahsulot yetarli emas.'}, status=status.HTTP_400_BAD_REQUEST)

    single_volume = sum(size.single_volume for size in product.sizes.all())
    total_amount = Decimal(quantity) * unit_price
    volume_m3 = Decimal(str(single_volume)) * Decimal(quantity)

    try:
        occurred_at = parse_occurred_at(request.data.get('occurred_at'))
    except ValueError:
        return Response({'error': 'Invalid occurred_at datetime format.'}, status=status.HTTP_400_BAD_REQUEST)

    debt = DebtRecord.objects.create(
        product=product,
        customer_fio=validated['customer_fio'],
        phone=phone,
        quantity=quantity,
        unit_price=unit_price,
        total_amount=total_amount,
        volume_m3=volume_m3,
        occurred_at=occurred_at,
    )

    product_cost = product.purchase_price
    product.quantity -= quantity
    product.sold_quantity += quantity
    product.save(update_fields=['quantity', 'sold_quantity', 'updated_at'])
    InventoryTransaction.objects.create(
        product=product,
        transaction_type=InventoryTransaction.TransactionType.OUT,
        quantity=quantity,
        unit_price=unit_price,
        unit_cost=product_cost,
        occurred_at=occurred_at,
    )

    return Response(DebtRecordSerializer(debt).data, status=status.HTTP_201_CREATED)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def debt_detail_view(request, pk: int):
    try:
        record = DebtRecord.objects.get(pk=pk)
    except DebtRecord.DoesNotExist:
        return Response({'error': 'Debt record not found.'}, status=status.HTTP_404_NOT_FOUND)

    record.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_debts_excel_view(request):
    records = DebtRecord.objects.select_related('product').order_by('-occurred_at')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Qarzdorlar'
    ws.append(['FIO', 'Mahsulot', 'Telefon', 'Soni', 'Bir dona narxi', 'Jami summa', 'Hajm (m³)', 'Sana', 'Soat'])
    for debt in records:
        local_dt = timezone.localtime(debt.occurred_at)
        ws.append([
            debt.customer_fio,
            debt.product.name,
            debt.phone,
            debt.quantity,
            float(debt.unit_price),
            float(debt.total_amount),
            float(debt.volume_m3),
            local_dt.strftime('%Y-%m-%d'),
            local_dt.strftime('%H:%M:%S'),
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return FileResponse(
        buffer,
        as_attachment=True,
        filename='qarzdorlar_royxati.xlsx',
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@api_view(['POST'])
@permission_classes([AllowAny])
def login_view(request):
    """User login view - returns JWT tokens"""
    username = request.data.get('username')
    password = request.data.get('password')
    
    if not username or not password:
        return Response(
            {'error': 'Username and password are required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    user = authenticate(username=username, password=password)
    
    if user is not None:
        refresh = RefreshToken.for_user(user)
        return Response({
            'token': str(refresh.access_token),
            'user': {
                'username': user.username,
                'role': 'admin' if user.is_superuser else 'user',
                'is_superuser': user.is_superuser
            }
        })
    else:
        return Response(
            {'error': 'Invalid credentials'},
            status=status.HTTP_401_UNAUTHORIZED
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def refresh_token_view(request):
    """Refresh JWT token"""
    refresh_token = request.data.get('refresh')
    
    if not refresh_token:
        return Response(
            {'error': 'Refresh token is required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        refresh = RefreshToken(refresh_token)
        return Response({
            'token': str(refresh.access_token)
        })
    except Exception:
        return Response(
            {'error': 'Invalid refresh token'},
            status=status.HTTP_401_UNAUTHORIZED
        )
